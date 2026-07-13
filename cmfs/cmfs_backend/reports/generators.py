"""
FILE: cmfs/cmfs_backend/reports/generators.py
ACTION: CREATE (Phase 10)

Aggregates budget/payments/gate data per ConventionUnit (and overall), then
renders Excel (OpenPyXL) and PDF (ReportLab) files to /media/reports/... and
records one `Report` row per file.

Colour coding follows the Financial Statement Colour Reference in the master
prompt:
  Income rows            -> #BDD7EE (blue)
  Overspent expense rows  -> #FFCCCC (red)
  Saving expense rows     -> #CCFFCC (green)
  Unbudgeted expense rows -> #FFE4B5 (orange)
  Section header rows     -> #D9D9D9 (bold, light grey)
  Summary surplus row     -> #CCFFCC (green)
  Summary deficit row     -> #FFCCCC (red)
  Written-off rows        -> #FFFF99 (yellow)
"""

import logging
import os
from io import BytesIO

from django.conf import settings
from django.utils import timezone

logger = logging.getLogger(__name__)

# ── Colours ──────────────────────────────────────────────────────────────────
BLUE = 'BDD7EE'
RED = 'FFCCCC'
GREEN = 'CCFFCC'
ORANGE = 'FFE4B5'
GREY = 'D9D9D9'
YELLOW = 'FFFF99'

# Colours matching the reference 'LITEIN SCHOOL FINANCIAL STATEMENT.xlsx'
# 'Income and Statement' sheet exactly, as specified by the user:
HEADER_ORANGE = 'FFC000'   # title block / block headers / category column-header rows
EXPENSE_GREEN = 'A9D08E'   # every expense line-item row (both blocks)
TOTAL_YELLOW = 'FFFF00'    # Total Income row + category EX-n subtotal rows
FOOD_BLUE = 'D9E1F2'       # FOOD category's subtotal (Total/Budgeted) cells
ACCOM_BLUE = '8EA9DB'      # ACCOM category on the BUDGETED VS ACTUALS side


# ── Data aggregation ─────────────────────────────────────────────────────────

def _counties_for_unit(unit):
    from conventions.models import County
    if unit is None:
        return County.objects.all()
    if unit.county_id:
        return County.objects.filter(pk=unit.county_id)
    if unit.region_id:
        return County.objects.filter(region_id=unit.region_id)
    return County.objects.all()


def _unit_context_labels(unit):
    """
    Describes where a ConventionUnit sits in the geography hierarchy, in
    both directions, for the overall report's per-unit overview table:
      - county-scope unit  -> its own county name, PLUS the region that
                               county belongs to.
      - regional-scope unit -> the region name, PLUS the counties that
                               make up that region.
      - national-scope unit -> no parent/children to show.
    Returns (primary_label, parent_label).
    """
    if unit.scope_type == 'county' and unit.county_id:
        county = unit.county
        region_name = county.region.name if county.region_id else 'Unknown Region'
        return county.name, f'{region_name} Region'

    if unit.scope_type == 'regional' and unit.region_id:
        from conventions.models import County
        region_name = unit.region.name
        county_names = list(
            County.objects.filter(region_id=unit.region_id).order_by('name').values_list('name', flat=True)
        )
        if county_names:
            parent = f"{len(county_names)} counties: " + ', '.join(county_names)
        else:
            parent = 'No counties assigned'
        return region_name, parent

    return 'National', None


def _build_unit_breakdown_and_insights(unit_data_pairs):
    """
    Builds the "OVERALL SUMMARY" report's real substance: a per-unit
    overview (which county/region each unit covers, and that unit's own
    income/expenditure/net balance) plus cross-unit comparative insights
    (which unit saved the most, which ran the biggest deficit, etc).

    `unit_data_pairs` is a list of (ConventionUnit, aggregate_data(...) or
    None) — None means that unit's own aggregation failed and it's skipped
    here (its own per-unit report will show as FAILED separately).
    """
    rows = []
    for unit, data in unit_data_pairs:
        if data is None:
            continue
        primary_label, parent_label = _unit_context_labels(unit)
        budgeted_expenditure = sum((el['estimated'] for el in data['expenditure_lines']), 0)
        actual_expenditure = data['total_expenditure']
        rows.append({
            'unit_id': unit.id,
            'scope_type': unit.scope_type,
            'display_name': primary_label,
            'parent_label': parent_label,
            'total_income': data['total_income'],
            'total_expenditure': actual_expenditure,
            'net_balance': data['net_balance'],
            'budgeted_expenditure': budgeted_expenditure,
            # Positive = came in under budget ("saved"); negative = overspent.
            'budget_variance': budgeted_expenditure - actual_expenditure,
            'delegate_count': data['delegate_count'],
            'checked_in_count': data['checked_in_count'],
        })

    def _top(rows_, key, want_max=True):
        if not rows_:
            return None
        chosen = max(rows_, key=lambda r: r[key]) if want_max else min(rows_, key=lambda r: r[key])
        return {'display_name': chosen['display_name'], 'value': chosen[key]}

    insights = {
        # Financial performance (income vs expenditure), matching the
        # existing SURPLUS (green) / DEFICIT (red) convention used
        # elsewhere in these reports.
        'top_surplus_unit': _top(rows, 'net_balance', want_max=True),
        'top_deficit_unit': _top([r for r in rows if r['net_balance'] < 0], 'net_balance', want_max=False),
        'top_income_unit': _top(rows, 'total_income', want_max=True),
        'top_expenditure_unit': _top(rows, 'total_expenditure', want_max=True),
        # Spending discipline vs each unit's own budget.
        'best_budget_saver': _top([r for r in rows if r['budget_variance'] > 0], 'budget_variance', want_max=True),
        'worst_overspender': _top([r for r in rows if r['budget_variance'] < 0], 'budget_variance', want_max=False),
    }

    return rows, insights


def aggregate_data(convention, unit=None):
    """Returns a dict with every figure needed to render the report set for
    `unit` (or the overall convention when unit is None)."""
    from delegates.models import Delegate, WriteOff
    from payments.models import Payment
    from budget.models import BudgetIncome, BudgetExpenseItem, ActualExpense
    from gate.models import Attendance

    counties = _counties_for_unit(unit)
    delegates_qs = Delegate.objects.filter(convention=convention, county__in=counties)
    payments_qs = Payment.objects.filter(delegate__in=delegates_qs, status='confirmed')

    if unit is not None:
        budget_incomes = BudgetIncome.objects.filter(convention_unit=unit)
        budget_expenses = BudgetExpenseItem.objects.filter(convention_unit=unit)
    else:
        budget_incomes = BudgetIncome.objects.filter(convention_unit__convention=convention)
        budget_expenses = BudgetExpenseItem.objects.filter(convention_unit__convention=convention)

    actual_expenses = ActualExpense.objects.filter(budget_expense_item__in=budget_expenses)
    write_offs_qs = WriteOff.objects.filter(delegate__in=delegates_qs)

    delegate_income = sum((p.amount_paid for p in payments_qs), 0)
    total_expenditure = sum((a.actual_total for a in actual_expenses), 0)
    total_written_off = sum((w.amount_written_off for w in write_offs_qs), 0)

    # Income lines: budgeted category vs actual.
    #   student/kessat/associate — actual always derived live from confirmed
    #     Payment records (never manually entered, can't be overridden).
    #   offering/exhibition — not tied to any Delegate/Payment, so the actual
    #     is whatever was manually recorded on BudgetIncome.actual_total
    #     (see budget.views.BudgetIncomeActualView). Defaults to 0 until set.
    # Registration fee schedule (Cost-Unit/Price for delegate categories) is
    # set on the Convention itself at creation time — one fee per category,
    # applied to every delegate in that category regardless of unit.
    category_fee = {
        'student': convention.fee_student,
        'kessat': convention.fee_kessat,
        'associate': convention.fee_associate,
    }

    income_lines = []
    for bi in budget_incomes:
        if bi.category in ('student', 'kessat', 'associate'):
            matching_payments = [p for p in payments_qs if p.delegate.category == bi.category]
            actual = sum((p.amount_paid for p in matching_payments), 0)
            # Count/Qty = number of delegates in this category (matches the
            # INCOMES sheet's per-category count), not number of payments —
            # a delegate can have multiple partial payments.
            count = delegates_qs.filter(category=bi.category).count()
            unit_price = category_fee.get(bi.category)
        else:
            actual = bi.actual_total if bi.actual_total is not None else 0
            count = None
            unit_price = None
        income_lines.append({
            'category': bi.get_category_display(),
            'category_code': bi.category,
            'estimated': bi.estimated_total,
            'actual': actual,
            'count': count,
            'unit_price': unit_price,
        })

    # total_income/net_balance must include offering + exhibition actuals
    # (delegate_income alone only covers student/kessat/associate payments).
    other_income = sum((il['actual'] for il in income_lines if il['category_code'] in ('offering', 'exhibition')), 0)
    total_income = delegate_income + other_income
    net_balance = total_income - total_expenditure

    # Per-category cash/paybill(M-Pesa) split, for the INCOMES sheet — "PAYBILL"
    # is this org's term for M-Pesa, so payment_method == 'mpesa' populates it.
    income_summary_rows = []
    for il in income_lines:
        if il['category_code'] in ('student', 'kessat', 'associate'):
            matching = [p for p in payments_qs if p.delegate.category == il['category_code']]
            cash = sum((p.amount_paid for p in matching if p.payment_method == 'cash'), 0)
            paybill = sum((p.amount_paid for p in matching if p.payment_method == 'mpesa'), 0)
            count = len(matching)
        else:
            # offering/exhibition: no per-payment breakdown exists, so the
            # full manually-recorded actual is shown under CASH (the
            # practical collection method for these categories).
            cash = il['actual']
            paybill = 0
            count = None
        income_summary_rows.append({
            'category': il['category'],
            'count': count,
            'expected': il['estimated'],
            'actual': il['actual'],
            'deficit': il['estimated'] - il['actual'],
            'cash': cash,
            'paybill': paybill,
        })

    cash_payments = [p for p in payments_qs if p.payment_method == 'cash']
    daily_cash = {}
    for p in cash_payments:
        day_label = p.timestamp.strftime('%a').upper()
        daily_cash[day_label] = daily_cash.get(day_label, 0) + p.amount_paid
    daily_cash_rows = [{'day': d, 'cash': amt} for d, amt in daily_cash.items()]

    # Expenditure lines: budgeted vs actual per line item.
    expenditure_lines = []
    for item in budget_expenses:
        item_actuals = [a for a in actual_expenses if a.budget_expense_item_id == item.id]
        actual_total = sum((a.actual_total for a in item_actuals), 0)
        if item.is_unbudgeted:
            colour = ORANGE
            label = 'UNBUDGETED'
        elif actual_total > item.estimated_total:
            colour = RED
            label = 'OVERSPENT'
        elif actual_total < item.estimated_total:
            colour = GREEN
            label = 'SAVING'
        else:
            colour = None
            label = ''
        # Actual qty/price aren't uniquely defined when an item has multiple
        # ActualExpense entries (multiple vouchers against one line) — show
        # the most recent entry's qty/price as representative, alongside
        # the always-correct summed actual_total.
        last_actual = item_actuals[-1] if item_actuals else None
        expenditure_lines.append({
            'item_code': item.item_code,
            'item_name': item.item_name,
            'category': item.category,
            'unit': item.unit,
            'estimated_qty': item.estimated_qty,
            'unit_price': item.unit_price,
            'days': item.days,
            'actual_qty': last_actual.actual_qty if last_actual else None,
            'actual_unit_price': last_actual.actual_unit_price if last_actual else None,
            'estimated': item.estimated_total,
            'actual': actual_total,
            'colour': colour,
            'label': label,
        })

    # Group expenditure lines by category, in CATEGORY_CHOICES order, so the
    # report can render one section (with its own subtotal) per category —
    # matching the reference financial statement's grouped layout.
    category_labels = dict(BudgetExpenseItem.CATEGORY_CHOICES)
    category_groups = []
    for cat_code, cat_label in BudgetExpenseItem.CATEGORY_CHOICES:
        lines = [el for el in expenditure_lines if el['category'] == cat_code]
        if not lines:
            continue
        budgeted_subtotal = sum((el['estimated'] for el in lines), 0)
        actual_subtotal = sum((el['actual'] for el in lines), 0)
        category_groups.append({
            'category': cat_code,
            'category_label': cat_label,
            'lines': lines,
            'budgeted_subtotal': budgeted_subtotal,
            'actual_subtotal': actual_subtotal,
        })

    # Rollup for the three "named people" categories — Catering Staff,
    # Speaker Tokens, Workers & Appreciation — pulled from actual spend,
    # for visibility in the report. This is a display rollup only: these
    # amounts are already counted once each inside their own category
    # group's subtotal above, so do NOT add this into total_expenditure
    # again.
    from budget.services import PEOPLE_APPRECIATION_CATEGORIES
    people_appreciation_totals = {cat: 0 for cat in PEOPLE_APPRECIATION_CATEGORIES}
    for el in expenditure_lines:
        if el['category'] in people_appreciation_totals:
            people_appreciation_totals[el['category']] += el['actual']
    people_appreciation_totals['combined_total'] = sum(
        (people_appreciation_totals[c] for c in PEOPLE_APPRECIATION_CATEGORIES), 0
    )

    vouchers = [{
        'voucher_number': a.voucher_number,
        'item_name': a.budget_expense_item.item_name,
        'category': a.budget_expense_item.category,
        'actual_total': a.actual_total,
        'actual_qty': a.actual_qty,
        'actual_unit_price': a.actual_unit_price,
        'authorized_by': a.authorized_by,
        'received_by': a.received_by,
        'entered_by_name': a.entered_by_name,
        'timestamp': a.timestamp,
    } for a in actual_expenses.order_by('voucher_number')]

    delegates = list(delegates_qs.select_related('county'))
    delegate_rows = [{
        'delegate_id': d.delegate_id or 'PENDING',
        'full_name': d.full_name,
        'category': d.category,
        'county': d.county.name,
        'fee_amount': d.fee_amount,
        'total_paid': d.total_paid,
        'balance_owed': d.balance_owed,
        'payment_status': d.payment_status,
    } for d in delegates]

    attendance_qs = Attendance.objects.filter(delegate__in=delegates_qs)
    checked_in_count = attendance_qs.filter(checked_in=True).count()
    attendance_rows = [{
        'delegate_name': a.delegate.full_name,
        'checked_in': a.checked_in,
        'checked_in_at': a.checked_in_at,
        'gate_location': a.gate_location,
    } for a in attendance_qs.select_related('delegate')]

    outstanding_rows = [d for d in delegate_rows if d['payment_status'] in ('INCOMPLETE', 'NOT_PAID')]
    write_off_rows = [{
        'delegate_name': w.delegate.full_name,
        'amount_written_off': w.amount_written_off,
        'reason': w.reason,
        'written_off_by_name': w.written_off_by_name,
        'written_off_at': w.written_off_at,
    } for w in write_offs_qs.select_related('delegate')]

    return {
        'counties': [c.name for c in counties],
        'total_income': total_income,
        'total_expenditure': total_expenditure,
        'total_written_off': total_written_off,
        'net_balance': net_balance,
        'income_lines': income_lines,
        'income_summary_rows': income_summary_rows,
        'daily_cash_rows': daily_cash_rows,
        'expenditure_lines': expenditure_lines,
        'category_groups': category_groups,
        'people_appreciation_totals': people_appreciation_totals,
        'vouchers': vouchers,
        'delegate_rows': delegate_rows,
        'delegate_count': len(delegate_rows),
        'checked_in_count': checked_in_count,
        'attendance_rows': attendance_rows,
        'outstanding_rows': outstanding_rows,
        'write_off_rows': write_off_rows,
    }


# ── Excel (OpenPyXL) ─────────────────────────────────────────────────────────

def _build_income_statement_sheet(ws, convention, unit, scope_label, data, fill_row, bold):
    """
    Renders the twin-block 'Income and Statement' sheet:
      LEFT  block (cols B-H): INCOMES AND EXPENDITURE
            Code | Item | Units | Count/Qty | Cost-Unit/Price | Days | Total
      RIGHT block (cols L-O): BUDGETED VS ACTUALS
            Code | Item | Total (Budgeted) | Total (Actuals)
    separated by a 3-column blank gutter (I, J, K).

    Both blocks share row numbers so each line lines up horizontally with
    its budgeted-vs-actual comparison, matching the reference
    'LITEIN SCHOOL FINANCIAL STATEMENT.xlsx' workbook — both in the figures
    shown (LEFT['total'] on an expense line is always the ACTUAL amount
    spent, so the LEFT block's totals tie out exactly to the ACTUALS column
    of the RIGHT block) and in the colour/border scheme:
      - Title block, block headers ('INCOMES AND EXPENDITURE' /
        'BUDGETED VS ACTUALS'), and every 'Units | Quantity | Days | Total'
        category column-header row -> HEADER_ORANGE (FFC000)
      - Every expense line-item row (both blocks)                -> EXPENSE_GREEN (A9D08E)
      - Total Income row + each category's EX-n subtotal row     -> TOTAL_YELLOW (FFFF00)
      - FOOD category's own subtotal (Total/Budgeted) cells      -> FOOD_BLUE (D9E1F2)
      - ACCOM category, BUDGETED VS ACTUALS side only            -> ACCOM_BLUE (8EA9DB)
      - Thin borders around every populated cell in both blocks, with the
        3-column gutter (I-K) and blank spacer rows left unbordered/unfilled
        for visual separation, exactly as in the reference workbook.

    The title block (rows 1-3 left blank, content starting row 4) shows,
    for each block, three merged/centred lines in this order: convention
    name -> block category label (INCOMES AND EXPENDITURE / BUDGETED VS
    ACTUALS) -> convention date range.
    """
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

    # Two blocks separated by a 3-column blank gutter (I, J, K) instead of
    # the single-column gutter used previously, for wider visual separation.
    LEFT = {'code': 'B', 'item': 'C', 'unit': 'D', 'qty': 'E', 'price': 'F', 'days': 'G', 'total': 'H'}
    RIGHT = {'code': 'L', 'item': 'M', 'budgeted': 'N', 'actual': 'O'}
    BLOCK_COLS = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'L', 'M', 'N', 'O']
    LEFT_SPAN = (2, 8)    # columns B..H
    RIGHT_SPAN = (12, 15)  # columns L..O

    thin = Side(style='thin', color='000000')
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def put(col_letter, row, value, font_bold=False, colour=None):
        cell = ws[f'{col_letter}{row}']
        cell.value = value
        if font_bold:
            cell.font = Font(bold=True)
        return cell

    def num(v):
        return float(v) if v not in (None, '') else None

    def style_row(row_idx, colour=None, cols=None):
        """Apply fill (if given) + a thin border to every cell of the twin
        block on this row. Only touches the block columns (B-H, L-O) —
        the 3-column gutter (I, J, K) between the two blocks and any
        columns outside the blocks are deliberately left untouched, for
        visual separation between INCOMES AND EXPENDITURE and BUDGETED VS
        ACTUALS."""
        cols = cols or BLOCK_COLS
        fill = PatternFill('solid', fgColor=colour) if colour else None
        for col in cols:
            cell = ws[f'{col}{row_idx}']
            if fill:
                cell.fill = fill
            cell.border = cell_border

    def merge_title(row_idx, text, span, size=14, colour=HEADER_ORANGE):
        """Writes `text` centred across the given (start_col, end_col)
        column span on row_idx, merging the cells and applying the fill.
        Border colour matches the fill so the seam between this row and
        the adjacent title rows (same fill colour) is invisible — the
        three stacked title lines read as one continuous block."""
        start_col, end_col = span
        ws.merge_cells(start_row=row_idx, start_column=start_col, end_row=row_idx, end_column=end_col)
        cell = ws.cell(row=row_idx, column=start_col, value=text)
        cell.font = Font(bold=True, size=size)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        fill = PatternFill('solid', fgColor=colour) if colour else None
        blend_side = Side(style='thin', color=colour) if colour else thin
        blend_border = Border(left=blend_side, right=blend_side, top=blend_side, bottom=blend_side)
        for c in range(start_col, end_col + 1):
            cel = ws.cell(row=row_idx, column=c)
            if fill:
                cel.fill = fill
            cel.border = blend_border

    date_range = ''
    if getattr(convention, 'start_date', None) and getattr(convention, 'end_date', None):
        date_range = f'{convention.start_date} — {convention.end_date}'

    # Rows 1-3 are left intentionally blank; the title block starts at
    # row 4. Each line is merged and centred across its block's full
    # column span, in the order: convention name -> block category label
    # (INCOMES AND EXPENDITURE / BUDGETED VS ACTUALS) -> convention dates.
    row = 4

    merge_title(row, convention.name, LEFT_SPAN, size=14)
    merge_title(row, convention.name, RIGHT_SPAN, size=14)
    row += 1

    merge_title(row, 'INCOMES AND EXPENDITURE', LEFT_SPAN, size=12)
    merge_title(row, 'BUDGETED VS ACTUALS', RIGHT_SPAN, size=12)
    row += 1

    merge_title(row, date_range, LEFT_SPAN, size=11)
    merge_title(row, date_range, RIGHT_SPAN, size=11)
    row += 1

    put(RIGHT['budgeted'], row, 'BUDGETED', font_bold=True)
    put(RIGHT['actual'], row, 'ACTUALS', font_bold=True)
    style_row(row, HEADER_ORANGE)
    row += 1

    put(LEFT['code'], row, 'Code')
    put(LEFT['item'], row, 'Delegates/Item')
    put(LEFT['unit'], row, 'Units')
    put(LEFT['qty'], row, 'Count/Qty')
    put(LEFT['price'], row, 'Cost-Unit/Price')
    put(LEFT['days'], row, 'Days')
    put(LEFT['total'], row, 'Total')
    put(RIGHT['code'], row, 'Code')
    put(RIGHT['item'], row, 'Delegates/Item')
    put(RIGHT['budgeted'], row, 'Total (Budgeted)')
    put(RIGHT['actual'], row, 'Total (Actuals)')
    style_row(row, HEADER_ORANGE)
    for col in BLOCK_COLS:
        ws[f'{col}{row}'].font = Font(bold=True)
    row += 1

    ex_counter = 0

    def next_ex_code():
        nonlocal ex_counter
        ex_counter += 1
        return f'EX-{ex_counter}'

    # -- INCOME section ------------------------------------------------------
    for il in data['income_lines']:
        is_other = il['category'] in ('Offering', 'Exhibition')
        if is_other:
            continue
        put(LEFT['item'], row, il['category'])
        put(LEFT['qty'], row, il['count'])
        put(LEFT['price'], row, num(il['unit_price']))
        put(LEFT['total'], row, num(il['actual']))
        put(RIGHT['item'], row, il['category'])
        put(RIGHT['budgeted'], row, num(il['estimated']))
        put(RIGHT['actual'], row, num(il['actual']))
        style_row(row)
        row += 1

    income_subtotal_budgeted = sum((float(il['estimated']) for il in data['income_lines']
                                     if il['category'] not in ('Offering', 'Exhibition')), 0.0)
    income_subtotal_actual = sum((float(il['actual']) for il in data['income_lines']
                                    if il['category'] not in ('Offering', 'Exhibition')), 0.0)
    code = next_ex_code()
    put(LEFT['code'], row, code, font_bold=True)
    put(LEFT['total'], row, income_subtotal_actual, font_bold=True)
    put(RIGHT['code'], row, code, font_bold=True)
    put(RIGHT['budgeted'], row, income_subtotal_budgeted, font_bold=True)
    put(RIGHT['actual'], row, income_subtotal_actual, font_bold=True)
    style_row(row, TOTAL_YELLOW)
    row += 2

    put(LEFT['item'], row, 'Other Sources', font_bold=True)
    put(RIGHT['item'], row, 'Other Sources', font_bold=True)
    style_row(row, HEADER_ORANGE)
    row += 1
    for il in data['income_lines']:
        if il['category'] not in ('Offering', 'Exhibition'):
            continue
        put(LEFT['item'], row, il['category'])
        put(LEFT['qty'], row, il['count'])
        put(LEFT['price'], row, num(il['unit_price']))
        put(LEFT['total'], row, num(il['actual']))
        put(RIGHT['item'], row, il['category'])
        put(RIGHT['budgeted'], row, num(il['estimated']))
        put(RIGHT['actual'], row, num(il['actual']))
        style_row(row)
        row += 1
    row += 1

    code = next_ex_code()
    put(LEFT['code'], row, code, font_bold=True)
    put(LEFT['item'], row, 'Total Income', font_bold=True)
    put(LEFT['total'], row, float(data['total_income']), font_bold=True)
    put(RIGHT['code'], row, code, font_bold=True)
    put(RIGHT['item'], row, 'Total Income', font_bold=True)
    total_budgeted_income = sum((float(il['estimated']) for il in data['income_lines']), 0.0)
    put(RIGHT['budgeted'], row, total_budgeted_income, font_bold=True)
    put(RIGHT['actual'], row, float(data['total_income']), font_bold=True)
    style_row(row, TOTAL_YELLOW)
    row += 2

    # -- EXPENSE sections, grouped by category ------------------------------
    total_budgeted_expenses = 0.0
    for group in data['category_groups']:
        is_food = group['category'] == 'FOOD'
        is_accom = group['category'] == 'ACCOM'

        put(LEFT['item'], row, 'EXPENSES', font_bold=True)
        put(LEFT['unit'], row, 'Units', font_bold=True)
        put(LEFT['qty'], row, 'Quantity', font_bold=True)
        put(LEFT['days'], row, 'Days', font_bold=True)
        put(LEFT['total'], row, 'Total', font_bold=True)
        put(RIGHT['item'], row, group['category_label'], font_bold=True)
        put(RIGHT['budgeted'], row, 'Total', font_bold=True)
        put(RIGHT['actual'], row, 'Total', font_bold=True)
        style_row(row, HEADER_ORANGE)
        row += 1

        for i, el in enumerate(group['lines'], start=1):
            put(LEFT['code'], row, i)
            put(LEFT['item'], row, el['item_name'])
            put(LEFT['unit'], row, el['unit'] if el['unit'] else None)
            put(LEFT['qty'], row, num(el['actual_qty']))
            put(LEFT['price'], row, num(el['actual_unit_price']))
            put(LEFT['days'], row, el['days'])
            # Total column on the INCOMES AND EXPENDITURE side must be the
            # ACTUAL amount spent (not the budgeted estimate) so it ties out
            # to the ACTUALS column (M) on the BUDGETED VS ACTUALS side.
            put(LEFT['total'], row, float(el['actual']))
            put(RIGHT['code'], row, i)
            put(RIGHT['item'], row, el['item_name'])
            put(RIGHT['budgeted'], row, float(el['estimated']))
            put(RIGHT['actual'], row, float(el['actual']))
            # Expense line rows are always EXPENSE_GREEN on the left block;
            # on the right block ACCOM gets its own ACCOM_BLUE colour while
            # every other category (including FOOD) stays EXPENSE_GREEN.
            style_row(row, EXPENSE_GREEN, cols=['B', 'C', 'D', 'E', 'F', 'G', 'H'])
            style_row(row, ACCOM_BLUE if is_accom else EXPENSE_GREEN, cols=['L', 'M', 'N', 'O'])
            row += 1

        code = next_ex_code()
        put(LEFT['code'], row, code, font_bold=True)
        put(LEFT['total'], row, float(group['actual_subtotal']), font_bold=True)
        put(RIGHT['code'], row, code, font_bold=True)
        put(RIGHT['budgeted'], row, float(group['budgeted_subtotal']), font_bold=True)
        put(RIGHT['actual'], row, float(group['actual_subtotal']), font_bold=True)
        if is_food:
            # FOOD's own Total/Budgeted subtotal cells are FOOD_BLUE instead
            # of the usual TOTAL_YELLOW used by every other category.
            style_row(row, FOOD_BLUE)
        elif is_accom:
            style_row(row, TOTAL_YELLOW, cols=['B', 'C', 'D', 'E', 'F', 'G', 'H'])
            style_row(row, ACCOM_BLUE, cols=['L', 'M', 'N', 'O'])
        else:
            style_row(row, TOTAL_YELLOW)
        total_budgeted_expenses += float(group['budgeted_subtotal'])
        row += 1

    row += 1

    # -- Workers, Speakers & Appreciation rollup (visibility only) ---------
    pat = data['people_appreciation_totals']
    put(LEFT['item'], row, 'Workers, Speakers & Appreciation (combined)', font_bold=True)
    put(LEFT['total'], row, float(pat['combined_total']), font_bold=True)
    put(RIGHT['item'], row, 'Workers, Speakers & Appreciation (combined)', font_bold=True)
    put(RIGHT['actual'], row, float(pat['combined_total']), font_bold=True)
    style_row(row, HEADER_ORANGE)
    row += 2

    # -- UNBUDGETED EXPENSES -------------------------------------------------
    unbudgeted_lines = [el for group in data['category_groups'] for el in group['lines'] if el['label'] == 'UNBUDGETED']
    if unbudgeted_lines:
        put(LEFT['item'], row, 'UNBUDGETED EXPENSES', font_bold=True)
        style_row(row, ORANGE)
        row += 1
        unb_total = 0.0
        for el in unbudgeted_lines:
            put(LEFT['item'], row, el['item_name'])
            put(LEFT['total'], row, float(el['actual']))
            put(RIGHT['item'], row, el['item_name'])
            put(RIGHT['actual'], row, float(el['actual']))
            style_row(row, ORANGE)
            unb_total += float(el['actual'])
            row += 1
        put(LEFT['item'], row, 'Total Unbudgeted', font_bold=True)
        put(LEFT['total'], row, unb_total, font_bold=True)
        style_row(row, TOTAL_YELLOW, cols=['B', 'C', 'D', 'E', 'F', 'G', 'H'])
        row += 2

    # -- WRITTEN OFF ----------------------------------------------------------
    if data['write_off_rows']:
        put(LEFT['item'], row, 'WRITTEN OFF', font_bold=True)
        style_row(row, YELLOW, cols=['B', 'C', 'D', 'E', 'F', 'G', 'H'])
        row += 1
        put(LEFT['code'], row, 'Delegate')
        put(LEFT['item'], row, 'Amount')
        put(LEFT['unit'], row, 'Reason')
        put(LEFT['qty'], row, 'Officer')
        put(LEFT['price'], row, 'Date')
        style_row(row, HEADER_ORANGE, cols=['B', 'C', 'D', 'E', 'F', 'G', 'H'])
        row += 1
        for w in data['write_off_rows']:
            put(LEFT['code'], row, w['delegate_name'])
            put(LEFT['item'], row, float(w['amount_written_off']))
            put(LEFT['unit'], row, f"WRITTEN OFF — {w['reason']}")
            put(LEFT['qty'], row, w['written_off_by_name'])
            put(LEFT['price'], row, w['written_off_at'].strftime('%Y-%m-%d') if w['written_off_at'] else '')
            style_row(row, YELLOW, cols=['B', 'C', 'D', 'E', 'F', 'G', 'H'])
            row += 1
        row += 1

    # -- TOTAL EXPENSES / SURPLUS-DEFICIT (bottom of both blocks) -----------
    # NOTE: data['total_expenditure'] already sums actuals across every
    # BudgetExpenseItem for this unit regardless of is_unbudgeted, and the
    # category-group subtotals above already include unbudgeted items too
    # (they're grouped by category like anything else). Do NOT add
    # unbudgeted_lines actuals again here — that would double-count them.
    actual_total_expenses = float(data['total_expenditure'])
    put(LEFT['item'], row, 'Total Expenses', font_bold=True)
    put(LEFT['total'], row, actual_total_expenses, font_bold=True)
    put(RIGHT['item'], row, 'Total Expenses', font_bold=True)
    put(RIGHT['budgeted'], row, total_budgeted_expenses, font_bold=True)
    put(RIGHT['actual'], row, actual_total_expenses, font_bold=True)
    style_row(row, HEADER_ORANGE)
    row += 1

    actual_surplus = float(data['total_income']) - actual_total_expenses
    budgeted_surplus = total_budgeted_income - total_budgeted_expenses
    left_colour = GREEN if actual_surplus >= 0 else RED
    right_colour = GREEN if budgeted_surplus >= 0 else RED
    put(LEFT['item'], row, 'Surplus /Deficit', font_bold=True)
    put(LEFT['total'], row, actual_surplus, font_bold=True)
    put(RIGHT['item'], row, 'Surplus /Deficit', font_bold=True)
    put(RIGHT['budgeted'], row, budgeted_surplus, font_bold=True)
    put(RIGHT['actual'], row, actual_surplus, font_bold=True)
    style_row(row, left_colour, cols=['B', 'C', 'D', 'E', 'F', 'G', 'H'])
    style_row(row, right_colour, cols=['L', 'M', 'N', 'O'])
    row += 1

    # -- Column widths (cosmetic, matches reference) ------------------------
    widths = {'B': 8, 'C': 34, 'D': 12, 'E': 10, 'F': 12, 'G': 8, 'H': 14,
              'I': 3, 'J': 3, 'K': 3,
              'L': 8, 'M': 34, 'N': 14, 'O': 14}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def _build_payment_voucher_sheet(ws, convention, scope_label, data, fill_row, bold):
    """
    Renders the Expenses sheet — matches the reference 'LITEIN SCHOOL
    FINANCIAL STATEMENT.xlsx' workbook's 'Expenses' sheet
    ("EXPENDITURE & PAYMENT VOUCHER ACCOUNT"): title block, then vouchers
    grouped under their own category header (FOOD, ACCOM, SECAD, etc. —
    same categories/order as the Income and Statement sheet), each voucher
    shown as PV number | item | qty | unit price | amount, with a category
    subtotal row and a grand total at the bottom.

    Per feedback, the "Approved By" / "Authorized By" / "Entered By" /
    timestamp fields are not shown here — they exist only as internal
    audit trail columns on ActualExpense and add clutter the reference
    workbook doesn't have. If that audit detail is needed later it belongs
    on a separate dedicated sheet, not on this one.

    Today's data-entry flow assigns one voucher number per ActualExpense
    entry (see budget/views.py::_next_voucher_number), so each item below
    reliably maps 1:1 to a voucher.
    """
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

    thin = Side(style='thin', color='000000')
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    COLS = ['A', 'B', 'C', 'D', 'E', 'F']

    def style_row(row_idx, colour=None):
        fill = PatternFill('solid', fgColor=colour) if colour else None
        for col in COLS:
            cell = ws[f'{col}{row_idx}']
            if fill:
                cell.fill = fill
            cell.border = cell_border

    def merge_title(row_idx, text, size=14, colour=HEADER_ORANGE):
        """Border colour matches the fill so the seam between this row and
        the adjacent title rows (same fill colour) is invisible — the
        three stacked title lines read as one continuous filled block."""
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=6)
        cell = ws.cell(row=row_idx, column=1, value=text)
        cell.font = Font(bold=True, size=size)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        fill = PatternFill('solid', fgColor=colour) if colour else None
        blend_side = Side(style='thin', color=colour) if colour else thin
        blend_border = Border(left=blend_side, right=blend_side, top=blend_side, bottom=blend_side)
        for col in COLS:
            cel = ws[f'{col}{row_idx}']
            if fill:
                cel.fill = fill
            cel.border = blend_border

    row = 1
    merge_title(row, f'{convention.name} — {scope_label}', size=14)
    row += 1
    merge_title(row, 'EXPENDITURE & PAYMENT VOUCHER ACCOUNT', size=12)
    row += 1
    date_range = ''
    if getattr(convention, 'start_date', None) and getattr(convention, 'end_date', None):
        date_range = f'{convention.start_date} — {convention.end_date}'
    merge_title(row, date_range, size=11)
    row += 1
    # Blank spacer row: touch a cell (rather than ws.append([])) so
    # ws.max_row correctly advances — mixing ws.append() calls with the
    # direct row-indexed writes used for the title block above would
    # otherwise leave ws.max_row stale by one row, throwing off every
    # ws.append()-based row that follows (header, vouchers, totals).
    ws.cell(row=row, column=1)

    # Vouchers grouped by category, in the same order as the category
    # groups on the Income and Statement sheet.
    vouchers_by_category = {}
    for v in data['vouchers']:
        vouchers_by_category.setdefault(v['category'], []).append(v)

    def blank_row():
        """Inserts a blank spacer row by touching a cell directly rather
        than ws.append([]) — an empty-list append advances openpyxl's
        internal row pointer without creating a real cell, which leaves
        ws.max_row stale and throws off every subsequent `ws.max_row + 1`
        computation used for headers/totals below."""
        r = ws.max_row + 1
        ws.cell(row=r, column=1)

    header_row = ws.max_row + 1
    ws.append(['PV #', 'Item', 'Qty', 'Unit Price (KES)', 'Amount (KES)'])
    style_row(header_row, HEADER_ORANGE)
    for c in range(1, 6):
        ws.cell(row=header_row, column=c).font = bold
    blank_row()

    grand_total = 0.0
    for group in data['category_groups']:
        cat_code = group['category']
        items = vouchers_by_category.get(cat_code, [])
        if not items:
            continue

        cat_header_row = ws.max_row + 1
        ws.append([group['category_label'].upper()])
        style_row(cat_header_row, HEADER_ORANGE)
        ws.cell(row=cat_header_row, column=1).font = bold

        subtotal = 0.0
        for v in items:
            row_idx = ws.max_row + 1
            qty = float(v['actual_qty']) if v['actual_qty'] is not None else None
            unit_price = float(v['actual_unit_price']) if v['actual_unit_price'] is not None else None
            ws.append([v['voucher_number'], v['item_name'], qty, unit_price, float(v['actual_total'])])
            style_row(row_idx, EXPENSE_GREEN)
            subtotal += float(v['actual_total'])

        total_row = ws.max_row + 1
        ws.append(['', f"Total — {group['category_label']}", '', '', subtotal])
        style_row(total_row, TOTAL_YELLOW)
        ws.cell(row=total_row, column=2).font = bold
        ws.cell(row=total_row, column=5).font = bold
        grand_total += subtotal
        blank_row()

    total_row = ws.max_row + 1
    ws.append(['TOTAL EXPENDITURE', '', '', '', grand_total])
    style_row(total_row, HEADER_ORANGE)
    ws.cell(row=total_row, column=1).font = bold
    ws.cell(row=total_row, column=5).font = bold

    for col, width in zip('ABCDE', [10, 34, 10, 16, 16]):
        ws.column_dimensions[col].width = width


def _build_incomes_sheet(ws, convention, scope_label, data, fill_row, bold):
    """
    Renders the INCOMES sheet — matches the reference workbook's 'INCOMES'
    sheet: title block, then one row per income category (CATEGORY | COUNT |
    EXPECTED | ACTUAL | DEFICIT | CASH | PAYBILL) with a totals row, a daily
    CASH breakdown table (DAY | CASH), and OFFERING/EXHIBITION + TOTAL INCOMES
    at the bottom. "PAYBILL" is this org's term for M-Pesa.

    Our data model has no receipt-batch concept (Payment isn't tied to a
    receipt-number range), so rows are per category rather than per batch —
    everything else mirrors the reference exactly.

    Layout matches the Income and Statement sheet's conventions: column A
    is left blank as a margin (content starts at column B), rows 1-3 are
    left blank, and the title block (starting row 4) shows three merged/
    centred lines in order: convention name -> 'INCOME SUMMARY' -> date
    range. Every populated cell has a thin border; header/total rows use
    HEADER_ORANGE/TOTAL_YELLOW and data rows use EXPENSE_GREEN, matching
    the Income and Statement sheet's palette.
    """
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

    # Content is shifted one column right (starts at B) so column A stays
    # blank as a left margin, matching the Income and Statement sheet.
    COLS = ['B', 'C', 'D', 'E', 'F', 'G', 'H']
    START_COL, END_COL = 2, 8  # B..H

    thin = Side(style='thin', color='000000')
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def put(col_letter, row_idx, value, font_bold=False):
        cell = ws[f'{col_letter}{row_idx}']
        cell.value = value
        if font_bold:
            cell.font = Font(bold=True)
        return cell

    def style_row(row_idx, colour=None, cols=None):
        cols = cols or COLS
        fill = PatternFill('solid', fgColor=colour) if colour else None
        for col in cols:
            cell = ws[f'{col}{row_idx}']
            if fill:
                cell.fill = fill
            cell.border = cell_border

    def merge_title(row_idx, text, size=14, colour=HEADER_ORANGE):
        """Border colour matches the fill so the seam between this row and
        the adjacent title rows (same fill colour) is invisible."""
        ws.merge_cells(start_row=row_idx, start_column=START_COL, end_row=row_idx, end_column=END_COL)
        cell = ws.cell(row=row_idx, column=START_COL, value=text)
        cell.font = Font(bold=True, size=size)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        fill = PatternFill('solid', fgColor=colour) if colour else None
        blend_side = Side(style='thin', color=colour) if colour else thin
        blend_border = Border(left=blend_side, right=blend_side, top=blend_side, bottom=blend_side)
        for col in COLS:
            cel = ws[f'{col}{row_idx}']
            if fill:
                cel.fill = fill
            cel.border = blend_border

    date_range = ''
    if getattr(convention, 'start_date', None) and getattr(convention, 'end_date', None):
        date_range = f'{convention.start_date} — {convention.end_date}'

    # Rows 1-3 left blank; title block starts at row 4, same merged 3-line
    # pattern as the Income and Statement sheet: convention name -> block
    # category label ('INCOME SUMMARY') -> convention date range.
    row = 4
    merge_title(row, convention.name, size=14)
    row += 1
    merge_title(row, 'INCOME SUMMARY', size=12)
    row += 1
    merge_title(row, date_range, size=11)
    row += 1
    row += 1  # blank spacer row

    header_row = row
    headers = ['Category', 'Count', 'Expected (KES)', 'Actual (KES)', 'Deficit (KES)', 'Cash (KES)', 'Paybill (KES)']
    for col, h in zip(COLS, headers):
        put(col, header_row, h, font_bold=True)
    style_row(header_row, HEADER_ORANGE)
    row += 1

    delegate_rows = [r for r in data['income_summary_rows'] if r['count'] is not None]
    other_rows = [r for r in data['income_summary_rows'] if r['count'] is None]

    for r in delegate_rows:
        vals = [r['category'], r['count'], float(r['expected']), float(r['actual']),
                float(r['deficit']), float(r['cash']), float(r['paybill'])]
        for col, v in zip(COLS, vals):
            put(col, row, v)
        style_row(row, EXPENSE_GREEN)
        row += 1

    total_row_idx = row
    vals = [
        'TOTAL', sum(r['count'] for r in delegate_rows),
        sum(float(r['expected']) for r in delegate_rows), sum(float(r['actual']) for r in delegate_rows),
        sum(float(r['deficit']) for r in delegate_rows), sum(float(r['cash']) for r in delegate_rows),
        sum(float(r['paybill']) for r in delegate_rows),
    ]
    for col, v in zip(COLS, vals):
        put(col, total_row_idx, v, font_bold=True)
    style_row(total_row_idx, TOTAL_YELLOW)
    row += 1
    row += 1  # blank spacer row

    # -- Daily cash breakdown ------------------------------------------------
    day_header_row = row
    put('B', day_header_row, 'DAY', font_bold=True)
    put('C', day_header_row, 'CASH (KES)', font_bold=True)
    style_row(day_header_row, HEADER_ORANGE, cols=['B', 'C'])
    row += 1

    daily_total = 0.0
    for r in data['daily_cash_rows']:
        put('B', row, r['day'])
        put('C', row, float(r['cash']))
        style_row(row, EXPENSE_GREEN, cols=['B', 'C'])
        daily_total += float(r['cash'])
        row += 1

    daily_total_row = row
    put('B', daily_total_row, 'TOTAL', font_bold=True)
    put('C', daily_total_row, daily_total, font_bold=True)
    style_row(daily_total_row, TOTAL_YELLOW, cols=['B', 'C'])
    row += 1
    row += 1  # blank spacer row

    # -- Offering / Exhibition + grand total ---------------------------------
    for r in other_rows:
        vals = [r['category'].upper(), '', float(r['expected']), float(r['actual']), float(r['deficit']),
                float(r['cash']), float(r['paybill'])]
        for col, v in zip(COLS, vals):
            put(col, row, v)
        style_row(row, EXPENSE_GREEN)
        row += 1

    grand_total_row = row
    grand_total_actual = float(data['total_income'])
    put('B', grand_total_row, 'TOTAL INCOMES', font_bold=True)
    put('E', grand_total_row, grand_total_actual, font_bold=True)
    style_row(grand_total_row, HEADER_ORANGE)

    ws.column_dimensions['A'].width = 4
    for col, width in zip('BCDEFGH', [16, 10, 16, 16, 16, 14, 14]):
        ws.column_dimensions[col].width = width


def _build_named_people_sheet(ws, data, category_code, fill_row, bold):
    """
    Lists every ActualExpense entry under a single category (STAFF, SPEAK,
    or APPR), showing who was paid (`received_by`), what for (`item_name`),
    how much, which voucher, and when — matching the reference workbook's
    dedicated named-list sheets ('litein workers', 'SPEAKERS', 'appreciation
    of convention work'). Uses `received_by` as the payee name since that
    field already captures who received the money for these categories.
    """
    from openpyxl.styles import Border, Side
    thin = Side(style='thin', color='000000')
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.append(['Name', 'Role / Reason', 'Amount (KES)', 'Voucher #', 'Authorized By', 'Date'])
    fill_row(ws, 1, HEADER_ORANGE)
    for col in range(1, 7):
        cell = ws.cell(row=1, column=col)
        cell.font = bold
        cell.border = cell_border

    lines = []
    for group in data['category_groups']:
        if group['category'] != category_code:
            continue
        for el in group['lines']:
            lines.append(el)

    matching_vouchers = [v for v in data['vouchers'] if any(
        v['item_name'] == el['item_name'] for el in lines
    )]

    total = 0.0
    for v in matching_vouchers:
        ws.append([v['received_by'], v['item_name'], float(v['actual_total']), v['voucher_number'],
                    v['authorized_by'], v['timestamp'].strftime('%Y-%m-%d') if v['timestamp'] else ''])
        row_idx = ws.max_row
        fill_row(ws, row_idx, EXPENSE_GREEN)
        for col in range(1, 7):
            ws.cell(row=row_idx, column=col).border = cell_border
        total += float(v['actual_total'])

    total_row = ws.max_row + 1
    fill_row(ws, total_row, TOTAL_YELLOW)
    for col in range(1, 7):
        ws.cell(row=total_row, column=col).border = cell_border
    ws.cell(row=total_row, column=2, value='TOTAL').font = bold
    ws.cell(row=total_row, column=3, value=total).font = bold

    for col, width in zip('ABCDEF', [22, 28, 16, 12, 18, 14]):
        ws.column_dimensions[col].width = width


def _build_overview_sheet(ws, convention, data, fill_row, bold):
    """
    Overall summary's headline sheet: one row per ConventionUnit (which
    county/region it covers, plus its own income/expenditure/net balance)
    and a comparative insights block (top saver, biggest deficit, etc).
    Only rendered for the overall report (convention_unit = NULL).
    """
    from openpyxl.styles import Font

    rows = data['unit_breakdown']
    insights = data['insights']

    ws.append([f'{convention.name} — Overall Summary'])
    fill_row(ws, ws.max_row, HEADER_ORANGE)
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=13)
    ws.append([f"Scope: {convention.get_scope_display()}  |  {convention.start_date} to {convention.end_date}"])
    ws.append([])

    # -- Grand totals across every unit -------------------------------------
    ws.append(['GRAND TOTALS (all units combined)'])
    fill_row(ws, ws.max_row, GREY)
    ws.cell(row=ws.max_row, column=1).font = bold
    ws.append(['Total Income (KES)', float(data['total_income'])])
    ws.append(['Total Expenditure (KES)', float(data['total_expenditure'])])
    net_row_label = 'SURPLUS' if data['net_balance'] >= 0 else 'DEFICIT'
    ws.append([net_row_label, float(data['net_balance'])])
    fill_row(ws, ws.max_row, GREEN if data['net_balance'] >= 0 else RED)
    ws.cell(row=ws.max_row, column=1).font = bold
    ws.append([])

    # -- Comparative insights ------------------------------------------------
    ws.append(['HIGHLIGHTS ACROSS UNITS'])
    fill_row(ws, ws.max_row, HEADER_ORANGE)
    ws.cell(row=ws.max_row, column=1).font = bold

    def insight_row(label, entry, colour=None, is_currency=True):
        if entry is None:
            ws.append([label, 'N/A', ''])
            return
        value = f"KES {float(entry['value']):,.2f}" if is_currency else entry['value']
        ws.append([label, entry['display_name'], value])
        if colour:
            fill_row(ws, ws.max_row, colour)

    insight_row('Highest Surplus (saved the most)', insights['top_surplus_unit'], GREEN)
    insight_row('Highest Deficit', insights['top_deficit_unit'], RED)
    insight_row('Highest Income Raised', insights['top_income_unit'], BLUE)
    insight_row('Highest Expenditure', insights['top_expenditure_unit'], ORANGE)
    insight_row('Best Budget Saver (most under budget)', insights['best_budget_saver'], GREEN)
    insight_row('Worst Overspender (most over budget)', insights['worst_overspender'], RED)
    ws.append([])

    # -- Per-unit breakdown table --------------------------------------------
    ws.append(['PER-UNIT BREAKDOWN'])
    fill_row(ws, ws.max_row, HEADER_ORANGE)
    ws.cell(row=ws.max_row, column=1).font = bold

    header = ['Unit', 'Covers', 'Delegates', 'Checked In', 'Income (KES)',
              'Expenditure (KES)', 'Net Balance (KES)', 'Budget Variance (KES)', 'Status']
    ws.append(header)
    fill_row(ws, ws.max_row, HEADER_ORANGE)
    for c in range(1, len(header) + 1):
        ws.cell(row=ws.max_row, column=c).font = bold

    for r in sorted(rows, key=lambda x: x['net_balance'], reverse=True):
        status = 'SURPLUS' if r['net_balance'] >= 0 else 'DEFICIT'
        ws.append([
            r['display_name'], r['parent_label'] or '—', r['delegate_count'], r['checked_in_count'],
            float(r['total_income']), float(r['total_expenditure']), float(r['net_balance']),
            float(r['budget_variance']), status,
        ])
        fill_row(ws, ws.max_row, GREEN if r['net_balance'] >= 0 else RED)

    if not rows:
        ws.append(['No units with generated data.'])

    for col, width in zip('ABCDEFGHI', [20, 40, 11, 11, 16, 18, 16, 18, 10]):
        ws.column_dimensions[col].width = width


def build_workbook(convention, unit, data) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    scope_label = unit.display_name if unit else 'Overall Summary'
    wb = Workbook()

    header_fill = PatternFill('solid', fgColor=GREY)
    bold = Font(bold=True)

    def fill_row(ws, row_idx, hex_colour):
        if not hex_colour:
            return
        fill = PatternFill('solid', fgColor=hex_colour)
        for cell in ws[row_idx]:
            cell.fill = fill

    # -- Overall summary report: Overview sheet ONLY ------------------------
    # The overall report exists purely for cross-unit comparison — each
    # unit's own detailed line items (income lines, vouchers, delegate
    # register, attendance, etc) already live in that unit's own report, so
    # repeating them here (pooled across every unit) added noise rather than
    # a summary. Per feedback: the overall report should show only the
    # general/comparison figures — mostly totals — for each unit.
    is_overall_summary = unit is None and data.get('unit_breakdown') is not None
    if is_overall_summary:
        ws_overview = wb.active
        ws_overview.title = 'Overview'
        _build_overview_sheet(ws_overview, convention, data, fill_row, bold)
        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    ws = wb.active

    # -- Sheet: Income and Statement (twin-block layout) ----------------
    # Matches the reference "LITEIN SCHOOL FINANCIAL STATEMENT.xlsx" —
    # "Income and Statement" sheet: an INCOMES AND EXPENDITURE block on the
    # left (cols B-H) and a BUDGETED VS ACTUALS block on the right (cols
    # J-M), row-aligned so each line item sits next to its budget/actual
    # comparison, with per-category subtotal rows and a Total Expenses /
    # Surplus-Deficit summary at the bottom of both blocks.
    ws.title = 'Income and Statement'
    _build_income_statement_sheet(ws, convention, unit, scope_label, data, fill_row, bold)

    # NOTE: There is no separate "Budget vs Actuals" sheet — that comparison
    # already lives inside the "Income and Statement" sheet's twin-block
    # layout (the BUDGETED VS ACTUALS block on the right), so a standalone
    # duplicate sheet was removed per feedback.

    # -- Sheet 2: Expenses (Payment Voucher Log) ---------------------------
    ws3 = wb.create_sheet('Expenses')
    _build_payment_voucher_sheet(ws3, convention, scope_label, data, fill_row, bold)

    # -- Sheet 3: INCOMES (income summary) ----------------------------------
    ws4 = wb.create_sheet('INCOMES')
    _build_incomes_sheet(ws4, convention, scope_label, data, fill_row, bold)

    # Shared helper for the remaining "simple table" sheets: borders every
    # populated cell and swaps the old flat grey (D9D9D9) header fill for
    # HEADER_ORANGE, matching the palette used on the Income and Statement
    # / INCOMES sheets.
    from openpyxl.styles import Border, Side
    thin = Side(style='thin', color='000000')
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def border_row(ws_, row_idx, n_cols):
        for col in range(1, n_cols + 1):
            ws_.cell(row=row_idx, column=col).border = cell_border

    # -- Sheet 5: Delegate Register ---------------------------------------
    ws5 = wb.create_sheet('Delegate Register')
    ws5.append(['Delegate ID', 'Name', 'Category', 'County', 'Fee (KES)', 'Paid (KES)', 'Balance (KES)', 'Status'])
    fill_row(ws5, 1, HEADER_ORANGE)
    for c in range(1, 9):
        ws5.cell(row=1, column=c).font = bold
    border_row(ws5, 1, 8)
    for d in data['delegate_rows']:
        ws5.append([d['delegate_id'], d['full_name'], d['category'], d['county'],
                    float(d['fee_amount']), float(d['total_paid']), float(d['balance_owed']), d['payment_status']])
        border_row(ws5, ws5.max_row, 8)
    for col, width in zip('ABCDEFGH', [14, 24, 12, 16, 12, 12, 14, 14]):
        ws5.column_dimensions[col].width = width

    # -- Sheet 6: Attendance ---------------------------------------
    ws6 = wb.create_sheet('Attendance')
    ws6.append(['Delegate', 'Checked In', 'Checked In At', 'Gate Location'])
    fill_row(ws6, 1, HEADER_ORANGE)
    for c in range(1, 5):
        ws6.cell(row=1, column=c).font = bold
    border_row(ws6, 1, 4)
    for a in data['attendance_rows']:
        ws6.append([a['delegate_name'], 'YES' if a['checked_in'] else 'NO',
                    a['checked_in_at'].strftime('%Y-%m-%d %H:%M') if a['checked_in_at'] else '', a['gate_location']])
        border_row(ws6, ws6.max_row, 4)
    ws6.append(['TOTAL CHECKED IN', data['checked_in_count'], '', ''])
    fill_row(ws6, ws6.max_row, TOTAL_YELLOW)
    border_row(ws6, ws6.max_row, 4)
    ws6.cell(row=ws6.max_row, column=1).font = bold
    for col, width in zip('ABCD', [24, 12, 18, 16]):
        ws6.column_dimensions[col].width = width

    # -- Sheet 7: Outstanding & Written-Off ---------------------------------------
    ws7 = wb.create_sheet('Outstanding & Written-Off')
    ws7.append(['OUTSTANDING PAYMENTS'])
    fill_row(ws7, ws7.max_row, HEADER_ORANGE)
    ws7.cell(row=ws7.max_row, column=1).font = bold
    border_row(ws7, ws7.max_row, 4)
    ws7.append(['Delegate ID', 'Name', 'Balance Owed (KES)', 'Status'])
    fill_row(ws7, ws7.max_row, HEADER_ORANGE)
    for c in range(1, 5):
        ws7.cell(row=ws7.max_row, column=c).font = bold
    border_row(ws7, ws7.max_row, 4)
    for d in data['outstanding_rows']:
        ws7.append([d['delegate_id'], d['full_name'], float(d['balance_owed']), d['payment_status']])
        border_row(ws7, ws7.max_row, 4)
    ws7.append([])
    ws7.append(['WRITTEN OFF'])
    fill_row(ws7, ws7.max_row, HEADER_ORANGE)
    ws7.cell(row=ws7.max_row, column=1).font = bold
    border_row(ws7, ws7.max_row, 5)
    ws7.append(['Delegate', 'Amount (KES)', 'Reason', 'Officer', 'Date'])
    fill_row(ws7, ws7.max_row, HEADER_ORANGE)
    for c in range(1, 6):
        ws7.cell(row=ws7.max_row, column=c).font = bold
    border_row(ws7, ws7.max_row, 5)
    for w in data['write_off_rows']:
        ws7.append([w['delegate_name'], float(w['amount_written_off']),
                    f"WRITTEN OFF — {w['reason']}", w['written_off_by_name'],
                    w['written_off_at'].strftime('%Y-%m-%d') if w['written_off_at'] else ''])
        fill_row(ws7, ws7.max_row, YELLOW)
        border_row(ws7, ws7.max_row, 5)
    for col, width in zip('ABCDE', [24, 16, 30, 16, 14]):
        ws7.column_dimensions[col].width = width

    # -- Sheet 8-10: named-people sheets (mirrors reference workbook's
    #    'litein workers' / 'SPEAKERS' / 'appreciation of convention work'
    #    sheets) ---------------------------------------------------------
    ws8 = wb.create_sheet('Catering & Security Staff')
    _build_named_people_sheet(ws8, data, 'STAFF', fill_row, bold)

    ws9 = wb.create_sheet('Speakers')
    _build_named_people_sheet(ws9, data, 'SPEAK', fill_row, bold)

    ws10 = wb.create_sheet('Workers & Appreciation')
    _build_named_people_sheet(ws10, data, 'APPR', fill_row, bold)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── PDF (ReportLab) ──────────────────────────────────────────────────────────

def build_pdf(convention, unit, data) -> bytes:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    scope_label = unit.display_name if unit else 'Overall Summary'
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    styles = getSampleStyleSheet()
    is_overall_summary = unit is None and data.get('unit_breakdown') is not None
    subtitle = 'Cross-Unit Comparison Summary' if is_overall_summary else 'Final Income &amp; Expenditure Statement'
    story = [
        Paragraph(f'{convention.name} — {scope_label}', styles['Title']),
        Paragraph(subtitle, styles['Heading2']),
        Spacer(1, 12),
    ]

    def hex_color(h):
        return colors.HexColor(f'#{h}')

    # -- Overall summary report: Overview section ONLY ----------------------
    # Cross-unit comparison is the whole point of the overall report — each
    # unit's own detailed line items (income lines, vouchers, delegate
    # register, attendance, etc) already live in that unit's own report.
    # Repeating them here (pooled across every unit) was noise, not a
    # summary, so the overall PDF stops after this section.
    if is_overall_summary:
        rows = data['unit_breakdown']
        insights = data['insights']

        story.append(Paragraph(f'Scope: {convention.get_scope_display()} &nbsp;|&nbsp; {convention.start_date} to {convention.end_date}', styles['Normal']))
        story.append(Spacer(1, 10))

        story.append(Paragraph('Highlights Across Units', styles['Heading2']))

        def fmt_insight(entry):
            if entry is None:
                return 'N/A'
            return f"{entry['display_name']} — KES {float(entry['value']):,.2f}"

        highlight_rows = [
            ['Highest Surplus (saved the most)', fmt_insight(insights['top_surplus_unit'])],
            ['Highest Deficit', fmt_insight(insights['top_deficit_unit'])],
            ['Highest Income Raised', fmt_insight(insights['top_income_unit'])],
            ['Highest Expenditure', fmt_insight(insights['top_expenditure_unit'])],
            ['Best Budget Saver (most under budget)', fmt_insight(insights['best_budget_saver'])],
            ['Worst Overspender (most over budget)', fmt_insight(insights['worst_overspender'])],
        ]
        highlight_table = Table(highlight_rows, colWidths=[8 * cm, 9 * cm])
        highlight_table.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 0), (0, -1), hex_color(GREY)),
        ]))
        story += [highlight_table, Spacer(1, 16)]

        story.append(Paragraph('Per-Unit Breakdown', styles['Heading2']))
        breakdown_data = [['Unit', 'Covers', 'Income (KES)', 'Expenditure (KES)', 'Net Balance (KES)', 'Status']]
        row_colours = []
        for r in sorted(rows, key=lambda x: x['net_balance'], reverse=True):
            status = 'SURPLUS' if r['net_balance'] >= 0 else 'DEFICIT'
            breakdown_data.append([
                r['display_name'], r['parent_label'] or '—',
                f"{float(r['total_income']):,.2f}", f"{float(r['total_expenditure']):,.2f}",
                f"{float(r['net_balance']):,.2f}", status,
            ])
            row_colours.append(GREEN if r['net_balance'] >= 0 else RED)
        if not rows:
            breakdown_data.append(['No units with generated data.', '', '', '', '', ''])

        breakdown_table = Table(breakdown_data, colWidths=[3.2 * cm, 6 * cm, 3 * cm, 3 * cm, 3 * cm, 2 * cm], repeatRows=1)
        style_cmds = [
            ('BACKGROUND', (0, 0), (-1, 0), hex_color(GREY)),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
        ]
        for i, colour in enumerate(row_colours, start=1):
            style_cmds.append(('BACKGROUND', (0, i), (-1, i), hex_color(colour)))
        breakdown_table.setStyle(TableStyle(style_cmds))
        story += [breakdown_table, Spacer(1, 16)]

        story.append(Paragraph('Grand Totals (all units combined)', styles['Heading2']))
        grand_rows = [
            ['Total Income', f"KES {float(data['total_income']):,.2f}"],
            ['Total Expenditure', f"KES {float(data['total_expenditure']):,.2f}"],
        ]
        grand_table = Table(grand_rows, colWidths=[8 * cm, 9 * cm])
        grand_table.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 0), (0, -1), hex_color(GREY)),
        ]))
        story += [grand_table, Spacer(1, 10)]

        summary_label = 'SURPLUS' if data['net_balance'] >= 0 else 'DEFICIT'
        summary_colour = GREEN if data['net_balance'] >= 0 else RED
        summary_table = Table([[summary_label, f"KES {float(data['net_balance']):,.2f}"]], colWidths=[8 * cm, 9 * cm])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), hex_color(summary_colour)),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        story += [summary_table]

        doc.build(story)
        return buf.getvalue()

    # Income table
    income_data = [['Category', 'Budgeted (KES)', 'Actual (KES)']]
    for il in data['income_lines']:
        income_data.append([il['category'], f"{float(il['estimated']):,.2f}", f"{float(il['actual']):,.2f}"])
    income_data.append(['TOTAL INCOME', '', f"{float(data['total_income']):,.2f}"])
    income_table = Table(income_data, colWidths=[7 * cm, 5 * cm, 5 * cm])
    style_cmds = [
        ('BACKGROUND', (0, 0), (-1, 0), hex_color(GREY)),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
    ]
    for i in range(1, len(income_data) - 1):
        style_cmds.append(('BACKGROUND', (0, i), (-1, i), hex_color(BLUE)))
    income_table.setStyle(TableStyle(style_cmds))
    story += [income_table, Spacer(1, 16)]

    # Expenditure tables — one per category group, each with its own
    # subtotal row, matching the grouped layout in the Excel export.
    story.append(Paragraph('Expenditure by Category', styles['Heading2']))
    for group in data['category_groups']:
        story.append(Paragraph(group['category_label'], styles['Heading3']))
        exp_data = [['Item', 'Budgeted (KES)', 'Actual (KES)', 'Status']]
        exp_colours = []
        for el in group['lines']:
            exp_data.append([el['item_name'], f"{float(el['estimated']):,.2f}", f"{float(el['actual']):,.2f}", el['label']])
            exp_colours.append(el['colour'])
        exp_data.append([
            f"Subtotal — {group['category_label']}", f"{float(group['budgeted_subtotal']):,.2f}",
            f"{float(group['actual_subtotal']):,.2f}", '',
        ])
        exp_table = Table(exp_data, colWidths=[7 * cm, 4 * cm, 4 * cm, 3 * cm])
        style_cmds = [
            ('BACKGROUND', (0, 0), (-1, 0), hex_color(GREY)),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BACKGROUND', (0, -1), (-1, -1), hex_color(GREY)),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ]
        for i, colour in enumerate(exp_colours, start=1):
            if colour:
                style_cmds.append(('BACKGROUND', (0, i), (-1, i), hex_color(colour)))
        exp_table.setStyle(TableStyle(style_cmds))
        story += [exp_table, Spacer(1, 10)]

    # Workers, Speakers & Appreciation rollup (visibility only — already
    # counted once inside the category subtotals above; not added twice
    # into any total below).
    pat = data['people_appreciation_totals']
    rollup_table = Table(
        [['Workers, Speakers & Appreciation (combined)', f"KES {float(pat['combined_total']):,.2f}"]],
        colWidths=[10 * cm, 6 * cm],
    )
    rollup_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    story += [rollup_table, Spacer(1, 16)]

    # Total expenditure — data['total_expenditure'] already sums actuals
    # across every BudgetExpenseItem (budgeted and unbudgeted alike), so
    # this is not recomputed from the category subtotals to avoid drift.
    total_table = Table(
        [['TOTAL EXPENDITURE', '', f"{float(data['total_expenditure']):,.2f}", '']],
        colWidths=[7 * cm, 4 * cm, 4 * cm, 3 * cm],
    )
    total_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), hex_color(GREY)),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    story += [total_table, Spacer(1, 16)]

    # Surplus / Deficit
    summary_label = 'SURPLUS' if data['net_balance'] >= 0 else 'DEFICIT'
    summary_colour = GREEN if data['net_balance'] >= 0 else RED
    summary_table = Table([[summary_label, f"KES {float(data['net_balance']):,.2f}"]], colWidths=[7 * cm, 8 * cm])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), hex_color(summary_colour)),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    story += [summary_table]

    doc.build(story)
    return buf.getvalue()


# ── Annual Summary (Phase 11) ────────────────────────────────────────────────
#
# Triggered automatically 7 days after a December convention reaches
# FINANCIALLY_CLOSED (see conventions.tasks.check_annual_summary_trigger),
# or on demand by Super Admin. Aggregates every FINANCIALLY_CLOSED
# convention whose end_date falls within the given calendar year. Contains:
# total delegates across every convention in the year, total income/
# expenses per convention + year total, surplus/deficit per convention per
# county, year-on-year comparison against the prior year's AnnualSummary
# (if one exists), top 10 counties by delegate count, payment collection
# efficiency per county, unbudgeted expense totals, and written-off amounts.

def aggregate_annual_summary(year: int) -> dict:
    from conventions.models import Convention, County
    from delegates.models import Delegate

    conventions = list(
        Convention.objects.filter(
            status=Convention.STATUS_FINANCIALLY_CLOSED,
            end_date__year=year,
        ).order_by('start_date')
    )

    per_convention = []
    total_delegates = 0
    total_income = 0
    total_expenditure = 0
    total_unbudgeted = 0
    total_written_off = 0

    # county_code -> accumulator, across every convention this year
    county_acc = {}

    def _county_bucket(county_name):
        return county_acc.setdefault(county_name, {
            'county': county_name, 'delegate_count': 0,
            'expected': 0, 'actual': 0, 'net_balance': 0,
        })

    for conv in conventions:
        try:
            data = aggregate_data(conv, unit=None)
        except Exception:
            logger.exception(f'aggregate_annual_summary: aggregation failed for convention {conv.id}')
            continue

        unbudgeted_total = sum(
            (el['actual'] for group in data['category_groups'] for el in group['lines'] if el['label'] == 'UNBUDGETED'),
            0,
        )

        per_convention.append({
            'convention_id': conv.id,
            'convention_name': conv.name,
            'scope': conv.scope,
            'start_date': conv.start_date,
            'end_date': conv.end_date,
            'delegate_count': data['delegate_count'],
            'total_income': data['total_income'],
            'total_expenditure': data['total_expenditure'],
            'net_balance': data['net_balance'],
            'unbudgeted_total': unbudgeted_total,
            'written_off_total': data['total_written_off'],
        })

        total_delegates += data['delegate_count']
        total_income += data['total_income']
        total_expenditure += data['total_expenditure']
        total_unbudgeted += unbudgeted_total
        total_written_off += data['total_written_off']

        # Per-county breakdown for this convention (surplus/deficit per
        # convention per county, and payment collection efficiency).
        for unit in conv.units.all():
            if unit.scope_type != 'county' or not unit.county_id:
                continue
            try:
                unit_data = aggregate_data(conv, unit)
            except Exception:
                continue
            bucket = _county_bucket(unit.county.name)
            bucket['delegate_count'] += unit_data['delegate_count']
            bucket['net_balance'] += unit_data['net_balance']
            expected = sum((il['estimated'] for il in unit_data['income_lines']), 0)
            bucket['expected'] += expected
            bucket['actual'] += unit_data['total_income']

    top_counties = sorted(county_acc.values(), key=lambda r: r['delegate_count'], reverse=True)[:10]

    collection_efficiency = []
    for c in county_acc.values():
        pct = (float(c['actual']) / float(c['expected']) * 100) if c['expected'] else None
        collection_efficiency.append({
            'county': c['county'], 'expected': c['expected'], 'actual': c['actual'],
            'efficiency_pct': pct,
        })
    collection_efficiency.sort(key=lambda r: (r['efficiency_pct'] is None, -(r['efficiency_pct'] or 0)))

    # Year-on-year comparison against the prior year's AnnualSummary, if any.
    from .models import AnnualSummary
    prior = AnnualSummary.objects.filter(year=year - 1, status='generated').first()
    year_on_year = None
    if prior and prior.summary_totals:
        prior_income = prior.summary_totals.get('total_income', 0)
        prior_expenditure = prior.summary_totals.get('total_expenditure', 0)
        prior_delegates = prior.summary_totals.get('total_delegates', 0)
        year_on_year = {
            'prior_year': year - 1,
            'income_change_pct': _pct_change(prior_income, total_income),
            'expenditure_change_pct': _pct_change(prior_expenditure, total_expenditure),
            'delegate_change_pct': _pct_change(prior_delegates, total_delegates),
        }

    return {
        'year': year,
        'conventions': per_convention,
        'total_delegates': total_delegates,
        'total_income': total_income,
        'total_expenditure': total_expenditure,
        'net_balance': total_income - total_expenditure,
        'total_unbudgeted': total_unbudgeted,
        'total_written_off': total_written_off,
        'county_breakdown': list(county_acc.values()),
        'top_counties': top_counties,
        'collection_efficiency': collection_efficiency,
        'year_on_year': year_on_year,
    }


def _pct_change(old, new):
    old = float(old or 0)
    new = float(new or 0)
    if old == 0:
        return None
    return (new - old) / old * 100.0


def build_annual_summary_workbook(data: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = 'Annual Summary'
    bold = Font(bold=True)

    def fill_row(ws_, row_idx, hex_colour):
        if not hex_colour:
            return
        fill = PatternFill('solid', fgColor=hex_colour)
        for cell in ws_[row_idx]:
            cell.fill = fill

    ws.append([f"KSCF Annual Summary — {data['year']}"])
    fill_row(ws, 1, HEADER_ORANGE)
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.append([])

    ws.append(['YEAR TOTALS'])
    fill_row(ws, ws.max_row, GREY)
    ws.cell(row=ws.max_row, column=1).font = bold
    ws.append(['Total Delegates (all conventions)', data['total_delegates']])
    ws.append(['Total Income (KES)', float(data['total_income'])])
    ws.append(['Total Expenditure (KES)', float(data['total_expenditure'])])
    label = 'SURPLUS' if data['net_balance'] >= 0 else 'DEFICIT'
    ws.append([label, float(data['net_balance'])])
    fill_row(ws, ws.max_row, GREEN if data['net_balance'] >= 0 else RED)
    ws.append(['Total Unbudgeted Expenses (KES)', float(data['total_unbudgeted'])])
    fill_row(ws, ws.max_row, ORANGE)
    ws.append(['Total Written-Off (KES)', float(data['total_written_off'])])
    fill_row(ws, ws.max_row, YELLOW)
    ws.append([])

    if data['year_on_year']:
        yoy = data['year_on_year']
        ws.append([f"YEAR-ON-YEAR vs {yoy['prior_year']}"])
        fill_row(ws, ws.max_row, GREY)
        ws.cell(row=ws.max_row, column=1).font = bold

        def fmt_pct(v):
            return f"{v:+.1f}%" if v is not None else 'N/A'

        ws.append(['Income change', fmt_pct(yoy['income_change_pct'])])
        ws.append(['Expenditure change', fmt_pct(yoy['expenditure_change_pct'])])
        ws.append(['Delegate count change', fmt_pct(yoy['delegate_change_pct'])])
        ws.append([])

    ws.append(['PER-CONVENTION BREAKDOWN'])
    fill_row(ws, ws.max_row, HEADER_ORANGE)
    ws.cell(row=ws.max_row, column=1).font = bold
    header = ['Convention', 'Scope', 'Start', 'End', 'Delegates', 'Income (KES)',
              'Expenditure (KES)', 'Net Balance (KES)', 'Unbudgeted (KES)', 'Written Off (KES)']
    ws.append(header)
    fill_row(ws, ws.max_row, HEADER_ORANGE)
    for c in range(1, len(header) + 1):
        ws.cell(row=ws.max_row, column=c).font = bold
    for conv in data['conventions']:
        ws.append([
            conv['convention_name'], conv['scope'], str(conv['start_date']), str(conv['end_date']),
            conv['delegate_count'], float(conv['total_income']), float(conv['total_expenditure']),
            float(conv['net_balance']), float(conv['unbudgeted_total']), float(conv['written_off_total']),
        ])
        fill_row(ws, ws.max_row, GREEN if conv['net_balance'] >= 0 else RED)
    ws.append([])

    ws.append(['TOP 10 COUNTIES BY DELEGATE COUNT'])
    fill_row(ws, ws.max_row, HEADER_ORANGE)
    ws.cell(row=ws.max_row, column=1).font = bold
    ws.append(['County', 'Delegates', 'Net Balance (KES)'])
    fill_row(ws, ws.max_row, HEADER_ORANGE)
    for c in range(1, 4):
        ws.cell(row=ws.max_row, column=c).font = bold
    for c in data['top_counties']:
        ws.append([c['county'], c['delegate_count'], float(c['net_balance'])])
    ws.append([])

    ws.append(['PAYMENT COLLECTION EFFICIENCY PER COUNTY'])
    fill_row(ws, ws.max_row, HEADER_ORANGE)
    ws.cell(row=ws.max_row, column=1).font = bold
    ws.append(['County', 'Expected (KES)', 'Actual (KES)', 'Efficiency %'])
    fill_row(ws, ws.max_row, HEADER_ORANGE)
    for c in range(1, 5):
        ws.cell(row=ws.max_row, column=c).font = bold
    for c in data['collection_efficiency']:
        pct = c['efficiency_pct']
        ws.append([c['county'], float(c['expected']), float(c['actual']),
                    f"{pct:.1f}%" if pct is not None else 'N/A'])
        if pct is not None:
            fill_row(ws, ws.max_row, GREEN if pct >= 90 else (ORANGE if pct >= 60 else RED))

    for col, width in zip('ABCDEFGHIJ', [24, 12, 12, 12, 12, 16, 16, 16, 16, 16]):
        ws.column_dimensions[col].width = width

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_annual_summary_pdf(data: dict) -> bytes:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    styles = getSampleStyleSheet()

    def hex_color(h):
        return colors.HexColor(f'#{h}')

    story = [
        Paragraph(f"KSCF Annual Summary — {data['year']}", styles['Title']),
        Spacer(1, 12),
    ]

    totals_rows = [
        ['Total Delegates', str(data['total_delegates'])],
        ['Total Income (KES)', f"{float(data['total_income']):,.2f}"],
        ['Total Expenditure (KES)', f"{float(data['total_expenditure']):,.2f}"],
        ['SURPLUS' if data['net_balance'] >= 0 else 'DEFICIT', f"{float(data['net_balance']):,.2f}"],
        ['Total Unbudgeted (KES)', f"{float(data['total_unbudgeted']):,.2f}"],
        ['Total Written-Off (KES)', f"{float(data['total_written_off']):,.2f}"],
    ]
    totals_table = Table(totals_rows, colWidths=[8 * cm, 8 * cm])
    totals_table.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('BACKGROUND', (0, 0), (0, -1), hex_color(GREY)),
    ]))
    story += [Paragraph('Year Totals', styles['Heading2']), totals_table, Spacer(1, 16)]

    if data['year_on_year']:
        yoy = data['year_on_year']

        def fmt_pct(v):
            return f"{v:+.1f}%" if v is not None else 'N/A'

        yoy_rows = [
            ['Income change', fmt_pct(yoy['income_change_pct'])],
            ['Expenditure change', fmt_pct(yoy['expenditure_change_pct'])],
            ['Delegate count change', fmt_pct(yoy['delegate_change_pct'])],
        ]
        yoy_table = Table(yoy_rows, colWidths=[8 * cm, 8 * cm])
        yoy_table.setStyle(TableStyle([('GRID', (0, 0), (-1, -1), 0.5, colors.grey)]))
        story += [Paragraph(f"Year-on-Year vs {yoy['prior_year']}", styles['Heading2']), yoy_table, Spacer(1, 16)]

    conv_data = [['Convention', 'Delegates', 'Income', 'Expenditure', 'Net']]
    for conv in data['conventions']:
        conv_data.append([
            conv['convention_name'], str(conv['delegate_count']),
            f"{float(conv['total_income']):,.0f}", f"{float(conv['total_expenditure']):,.0f}",
            f"{float(conv['net_balance']):,.0f}",
        ])
    conv_table = Table(conv_data, colWidths=[6 * cm, 2.5 * cm, 3 * cm, 3 * cm, 3 * cm], repeatRows=1)
    conv_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), hex_color(GREY)),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
    ]))
    story += [Paragraph('Per-Convention Breakdown', styles['Heading2']), conv_table, Spacer(1, 16)]

    top_data = [['County', 'Delegates']] + [[c['county'], str(c['delegate_count'])] for c in data['top_counties']]
    top_table = Table(top_data, colWidths=[8 * cm, 4 * cm], repeatRows=1)
    top_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), hex_color(GREY)),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    story += [Paragraph('Top 10 Counties by Delegate Count', styles['Heading2']), top_table, Spacer(1, 16)]

    eff_data = [['County', 'Expected', 'Actual', 'Efficiency %']]
    for c in data['collection_efficiency']:
        pct = c['efficiency_pct']
        eff_data.append([c['county'], f"{float(c['expected']):,.0f}", f"{float(c['actual']):,.0f}",
                          f"{pct:.1f}%" if pct is not None else 'N/A'])
    eff_table = Table(eff_data, colWidths=[6 * cm, 3.5 * cm, 3.5 * cm, 3 * cm], repeatRows=1)
    eff_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), hex_color(GREY)),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    story += [Paragraph('Payment Collection Efficiency per County', styles['Heading2']), eff_table]

    doc.build(story)
    return buf.getvalue()


def generate_annual_summary_files(year: int, triggered_by=None):
    """
    Aggregates and renders the annual summary (xlsx + pdf), saves them to
    /media/reports/annual/{year}/summary.{xlsx|pdf}, and records/updates
    the AnnualSummary row. Returns the AnnualSummary instance.
    """
    from .models import AnnualSummary

    summary, _ = AnnualSummary.objects.get_or_create(year=year)
    summary.triggered_by_id = triggered_by
    try:
        data = aggregate_annual_summary(year)
        xlsx_bytes = build_annual_summary_workbook(data)
        pdf_bytes = build_annual_summary_pdf(data)

        xlsx_rel = f'reports/annual/{year}/summary.xlsx'
        pdf_rel = f'reports/annual/{year}/summary.pdf'
        for rel, content in ((xlsx_rel, xlsx_bytes), (pdf_rel, pdf_bytes)):
            abs_path = os.path.join(settings.MEDIA_ROOT, rel)
            os.makedirs(os.path.dirname(abs_path), exist_ok=True)
            with open(abs_path, 'wb') as f:
                f.write(content)

        summary.xlsx_path = xlsx_rel
        summary.pdf_path = pdf_rel
        summary.summary_totals = {
            'total_income': float(data['total_income']),
            'total_expenditure': float(data['total_expenditure']),
            'total_delegates': data['total_delegates'],
        }
        summary.status = 'generated'
        summary.generated_at = timezone.now()
        summary.error_message = ''
        summary.save()
    except Exception as e:
        logger.exception(f'generate_annual_summary_files: failed for year={year}')
        summary.status = 'failed'
        summary.error_message = str(e)[:2000]
        summary.save()
    return summary


# ── Orchestration ────────────────────────────────────────────────────────────

def generate_reports_for_convention(convention, report_type, generated_by_id=None):
    """
    Generates the overall (convention_unit=NULL) report plus one report per
    ConventionUnit, in both xlsx and pdf. Returns the list of Report rows
    created (status generated/failed).

    Each unit's data is aggregated once and reused for both that unit's own
    report AND the overall report's per-unit breakdown/insights — so the
    overall report isn't just the pooled totals, it's a real summary of how
    each unit did relative to the others.
    """
    from .models import Report

    created = []
    units = list(convention.units.all())

    unit_data_pairs = []
    for unit in units:
        try:
            unit_data = aggregate_data(convention, unit)
        except Exception:
            logger.exception(f'report aggregation failed for unit={unit.id}')
            unit_data = None
        unit_data_pairs.append((unit, unit_data))

    unit_breakdown, insights = _build_unit_breakdown_and_insights(unit_data_pairs)
    overall_extra = {'unit_breakdown': unit_breakdown, 'insights': insights}

    created += _generate_single(convention, None, report_type, generated_by_id, Report, overall_extra=overall_extra)

    for unit, unit_data in unit_data_pairs:
        created += _generate_single(
            convention, unit, report_type, generated_by_id, Report, precomputed_data=unit_data,
        )
    return created


def _generate_single(convention, unit, report_type, generated_by_id, Report, precomputed_data=None, overall_extra=None):
    # Replace, don't accumulate: each (convention, unit, report_type,
    # format) slot should only ever have one row. update_or_create keys on
    # exactly that slot, so re-generating (the dev "Generate Reports Now"
    # button, a second financial-close retry, opening day re-triggered)
    # reuses the existing row instead of racing a delete-then-create,
    # which could leave duplicate rows if two generation calls overlapped.
    # Backed by Report.Meta's uniq_report_overall_slot / uniq_report_unit_slot
    # constraints as a DB-level guarantee.
    results = []
    if precomputed_data is not None:
        data = precomputed_data
    else:
        try:
            data = aggregate_data(convention, unit)
        except Exception as e:
            logger.exception('report aggregation failed')
            for fmt in ('xlsx', 'pdf'):
                report, _ = Report.objects.update_or_create(
                    convention=convention, convention_unit=unit, report_type=report_type, format=fmt,
                    defaults=dict(
                        status='failed', error_message=str(e)[:2000],
                        generated_by_id=generated_by_id, file_path='', generated_at=None,
                    ),
                )
                results.append(report)
            return results

    if overall_extra:
        data = {**data, **overall_extra}

    for fmt in ('xlsx', 'pdf'):
        report, _ = Report.objects.update_or_create(
            convention=convention, convention_unit=unit, report_type=report_type, format=fmt,
            defaults=dict(status='pending', generated_by_id=generated_by_id, error_message=''),
        )
        try:
            content = build_workbook(convention, unit, data) if fmt == 'xlsx' else build_pdf(convention, unit, data)
            unit_folder = str(unit.id) if unit else 'overall'
            rel_path = f'reports/{convention.id}/{unit_folder}/{report_type}.{fmt}'
            abs_path = os.path.join(settings.MEDIA_ROOT, rel_path)
            os.makedirs(os.path.dirname(abs_path), exist_ok=True)
            with open(abs_path, 'wb') as f:
                f.write(content)
            report.file_path = rel_path
            report.status = 'generated'
            report.generated_at = timezone.now()
            report.save()
        except Exception as e:
            logger.exception(f'report generation failed for convention={convention.id} unit={unit}')
            report.status = 'failed'
            report.error_message = str(e)[:2000]
            report.save()
        results.append(report)
    return results